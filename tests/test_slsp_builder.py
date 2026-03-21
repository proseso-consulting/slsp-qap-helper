# tests/test_slsp_builder.py
import io
import pytest
from openpyxl import load_workbook
from slsp_builder import (
    aggregate_by_tin,
    build_slsp_rows,
    write_slsp_xlsx,
    write_slsp_dat,
)


@pytest.fixture
def sample_bill_rows():
    return [
        {
            "tin": "123456789",
            "registered_name": "ACME CORP",
            "last_name": "",
            "first_name": "",
            "middle_name": "",
            "street": "123 MAIN ST",
            "city": "MAKATI",
            "date": "2026-01-15",
            "exempt_amount": 0,
            "zero_rated_amount": 0,
            "services_amount": 10000.00,
            "capital_goods_amount": 0,
            "other_goods_amount": 0,
            "input_tax": 1200.00,
            "source": "bill",
        },
    ]


@pytest.fixture
def sample_je_rows():
    return [
        {
            "tin": "987654321",
            "registered_name": "VENDOR TWO",
            "last_name": "TWO",
            "first_name": "VENDOR",
            "middle_name": "",
            "street": "456 SIDE ST",
            "city": "QUEZON CITY",
            "date": "2026-02-10",
            "exempt_amount": 0,
            "zero_rated_amount": 0,
            "services_amount": 0,
            "capital_goods_amount": 5000.00,
            "other_goods_amount": 0,
            "input_tax": 600.00,
            "source": "journal_entry",
        },
    ]


_FILING_COMPANY = {
    "tin": "999888777",
    "registered_name": "FILING CORP",
    "first_name": "", "middle_name": "", "last_name": "",
    "street": "1 MAIN ST", "city": "MANILA",
    "rdo": "042",
}


class TestAggregateByTin:
    def test_merges_two_rows_same_tin(self):
        rows = [
            {"tin": "111222333", "registered_name": "ACME", "last_name": "", "first_name": "",
             "middle_name": "", "street": "1ST", "city": "MNL", "date": "2026-01-15",
             "services_amount": 10000.0, "input_tax": 1200.0,
             "exempt_amount": 0, "zero_rated_amount": 0,
             "capital_goods_amount": 0, "other_goods_amount": 0,
             "taxable_amount": 0, "tax_amount": 0, "source": "bill"},
            {"tin": "111222333", "registered_name": "ACME", "last_name": "", "first_name": "",
             "middle_name": "", "street": "1ST", "city": "MNL", "date": "2026-01-28",
             "services_amount": 5000.0, "input_tax": 600.0,
             "exempt_amount": 0, "zero_rated_amount": 0,
             "capital_goods_amount": 0, "other_goods_amount": 0,
             "taxable_amount": 0, "tax_amount": 0, "source": "bill"},
        ]
        result = aggregate_by_tin(rows)
        assert len(result) == 1
        assert result[0]["services_amount"] == 15000.0
        assert result[0]["input_tax"] == 1800.0

    def test_keeps_distinct_tins_separate(self, sample_bill_rows, sample_je_rows):
        merged = build_slsp_rows(sample_bill_rows, sample_je_rows)
        result = aggregate_by_tin(merged)
        assert len(result) == 2

    def test_empty_input(self):
        assert aggregate_by_tin([]) == []


class TestBuildSlspRows:
    def test_merges_bills_and_jes(self, sample_bill_rows, sample_je_rows):
        merged = build_slsp_rows(sample_bill_rows, sample_je_rows)
        assert len(merged) == 2

    def test_sorted_by_date(self, sample_bill_rows, sample_je_rows):
        merged = build_slsp_rows(sample_bill_rows, sample_je_rows)
        dates = [r["date"] for r in merged]
        assert dates == sorted(dates)

    def test_empty_jes(self, sample_bill_rows):
        merged = build_slsp_rows(sample_bill_rows, [])
        assert len(merged) == 1

    def test_empty_bills_and_jes(self):
        merged = build_slsp_rows([], [])
        assert merged == []


class TestWriteSlspXlsx:
    def test_produces_valid_xlsx(self, sample_bill_rows, sample_je_rows):
        merged = build_slsp_rows(sample_bill_rows, sample_je_rows)
        buf = io.BytesIO()
        write_slsp_xlsx(merged, buf, report_type="purchases")
        buf.seek(0)
        wb = load_workbook(buf)
        ws = wb.active
        assert ws.max_row == 3  # header + 2 data rows
        headers = [cell.value for cell in ws[1]]
        assert "TIN" in headers
        assert "Input Tax" in headers
        assert "Source" in headers


class TestWriteSlspDat:
    def test_produces_dat_lines(self, sample_bill_rows):
        dat = write_slsp_dat(sample_bill_rows, report_type="purchases", filing_tin="999888777")
        assert "D,P," in dat
        assert "123456789" in dat
        assert "10000.00" in dat

    def test_empty_rows_returns_empty_string(self):
        dat = write_slsp_dat([], report_type="purchases", filing_tin="999888777")
        assert dat == ""

    def test_h_record_prepended_when_company_and_period_given(self, sample_bill_rows):
        dat = write_slsp_dat(
            sample_bill_rows, report_type="purchases",
            filing_tin="999888777", period_end="2026-01-31",
            company=_FILING_COMPANY,
        )
        lines = [l for l in dat.splitlines() if l]
        assert lines[0].startswith("H,P,")
        assert lines[1].startswith("D,P,")

    def test_no_h_record_without_company(self, sample_bill_rows):
        dat = write_slsp_dat(
            sample_bill_rows, report_type="purchases",
            filing_tin="999888777", period_end="2026-01-31",
        )
        assert not dat.startswith("H,")

    def test_period_end_overrides_row_dates(self, sample_bill_rows):
        dat = write_slsp_dat(
            sample_bill_rows, report_type="purchases",
            filing_tin="999888777", period_end="2026-01-31",
        )
        assert "01/31/2026" in dat
        assert "01/15/2026" not in dat  # original row date should not appear
