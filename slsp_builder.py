# slsp_builder.py
"""SLSP report builder — merges bill and JE rows, outputs XLSX or DAT.

SLSP = Summary List of Sales/Purchases.
SLP = Summary List of Purchases (in_invoice, in_refund)
SLS = Summary List of Sales (out_invoice, out_refund)
"""

from __future__ import annotations

from typing import BinaryIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from bir_format import fmt_date_slsp, slp_dat_header, slp_dat_line, sls_dat_header, sls_dat_line

SLP_COLUMNS = [
    "TIN",
    "Registered Name",
    "Address",
    "Exempt Amount",
    "Zero-Rated Amount",
    "Services Amount",
    "Capital Goods Amount",
    "Other Goods Amount",
    "Input Tax",
    "Source",
]

SLS_COLUMNS = [
    "TIN",
    "Registered Name",
    "Address",
    "Exempt Amount",
    "Zero-Rated Amount",
    "Taxable Amount",
    "Output Tax",
    "Source",
]


_AMOUNT_FIELDS = frozenset(
    [
        "exempt_amount",
        "zero_rated_amount",
        "services_amount",
        "capital_goods_amount",
        "other_goods_amount",
        "input_tax",
        "taxable_amount",
        "tax_amount",
    ]
)


def aggregate_by_tin(rows: list[dict]) -> list[dict]:
    """Collapse rows to one entry per TIN, summing all amount fields.

    Non-amount fields (name, address, etc.) are taken from the first occurrence
    of each TIN. This mirrors the SUMIFS-per-TIN logic in the BIR spreadsheet
    — the DAT format requires one D record per counterparty, not one per invoice.
    """
    groups: dict[str, dict] = {}
    for row in rows:
        tin = row["tin"]
        if tin not in groups:
            groups[tin] = {**row, **{f: 0.0 for f in _AMOUNT_FIELDS}}
        for f in _AMOUNT_FIELDS:
            groups[tin][f] = round(groups[tin][f] + row.get(f, 0.0), 2)
    return list(groups.values())


def build_slsp_rows(bill_rows: list[dict], je_rows: list[dict]) -> list[dict]:
    """Merge bill and JE rows, sorted by date ascending."""
    merged = list(bill_rows) + list(je_rows)
    return sorted(merged, key=lambda r: r.get("date", ""))


def write_slsp_xlsx(rows: list[dict], output: BinaryIO, report_type: str = "purchases") -> None:
    """Write SLSP rows to an XLSX file in the given output buffer."""
    wb = Workbook()
    ws = wb.active
    ws.title = "SLP" if report_type == "purchases" else "SLS"

    columns = SLP_COLUMNS if report_type == "purchases" else SLS_COLUMNS
    header_fill = PatternFill(start_color="293750", end_color="293750", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        if report_type == "purchases":
            values = [
                row.get("tin", ""),
                row.get("registered_name", ""),
                f"{row.get('street', '')} {row.get('city', '')}".strip(),
                row.get("exempt_amount", 0),
                row.get("zero_rated_amount", 0),
                row.get("services_amount", 0),
                row.get("capital_goods_amount", 0),
                row.get("other_goods_amount", 0),
                row.get("input_tax", 0),
                row.get("source", ""),
            ]
        else:
            values = [
                row.get("tin", ""),
                row.get("registered_name", ""),
                f"{row.get('street', '')} {row.get('city', '')}".strip(),
                row.get("exempt_amount", 0),
                row.get("zero_rated_amount", 0),
                row.get("taxable_amount", 0),
                row.get("tax_amount", 0),
                row.get("source", ""),
            ]
        for col_idx, value in enumerate(values, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    wb.save(output)


def write_slsp_dat(
    rows: list[dict],
    report_type: str = "purchases",
    filing_tin: str = "000000000",
    period_end: str = "",
    company: dict | None = None,
) -> str:
    """Build DAT file content as a string.

    If company is supplied, prepends an H record with filing-entity details and
    period totals — required for BIR-valid DAT files.
    period_end: YYYY-MM-DD (e.g. "2025-11-30"). When provided, every D row uses
    this as its date (per BIR spec — D rows carry the period end date, not
    individual transaction dates).
    """
    if not rows:
        return ""

    period_fmt = fmt_date_slsp(period_end) if period_end else ""
    line_fn = slp_dat_line if report_type == "purchases" else sls_dat_line
    header_fn = slp_dat_header if report_type == "purchases" else sls_dat_header

    lines = []
    if company and period_fmt:
        lines.append(header_fn(company, rows, period_fmt))

    for row in rows:
        row_date = period_fmt if period_fmt else fmt_date_slsp(row["date"])
        dat_row = {**row, "date": row_date}
        lines.append(line_fn(dat_row, filing_tin=filing_tin))

    return "\r\n".join(lines) + "\r\n"
